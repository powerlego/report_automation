import os
import time


from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from dotenv import load_dotenv
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from data_grabber import DataGrabber


def main():
   
    wb = load_workbook(filename= os.path.join(os.environ['USERPROFILE'], 'Desktop', 'WTD Tracker Menards 2023.xlsx'))
    print(wb.sheetnames)
    # config = load_dotenv(".env")
    # data_grabber = DataGrabber(os.environ["PASSWORD1"])
    # driver = data_grabber.driver
    # print(data_grabber.base_url)
    # driver.get(data_grabber.base_url)
    # WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="contentFrame"]')))
    # driver.switch_to.frame(driver.find_element(By.XPATH, '//*[@id="contentFrame"]'))
    # element = driver.find_element(By.ID, "username")
    # element.send_keys(data_grabber.username1)
    # element = driver.find_element(By.ID, "passcode")
    # element.send_keys(data_grabber.password1)
    # time.sleep(200)
    
    # driver.quit()


if __name__ == "__main__":
    main()
